# Importing the necessary packages
import streamlit as st
import openpyxl 
import pandas as pd 
from itertools import islice
import datetime
import os 
import glob
import re
from difflib import get_close_matches
from streamlit_option_menu import option_menu


#Functions
# Start row for all CRTA (CRT & Extended Role CRT)
def start_row(worksheet):
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value == "Extended Role CRT":
                start_row = cell.row
                
    return start_row

# All CRT & Extended Role CRT namelist and cell id
def crt_names(worksheet):
    crt_names = {}
    for row in worksheet.iter_rows(min_row=(start_row(worksheet))+1, min_col=0, max_col=1 ):
        for cell in row:
            if isinstance(cell.value, str):
                crt_names[cell.value] = cell 
    return crt_names

# Converting the time to datetime format 
def time_format(time):
    hours = int(time)
    minutes = int((time - hours) * 100)
    
    time_obj = datetime.time(hour=hours, minute=minutes)
    
    return time_obj

# Finding the total hours, if shift is more than 7 hours, a mandatory 1 hour break is given 
def time_in_hours(timing):
    list_hour = []
    for i in range(0,len(timing),2):
        time1 = timing[i]
        time2 = timing[i+1]
        
        if time2 < time1:
            dt1 = datetime.datetime.combine(datetime.date.today(), time1)
            dt2 = datetime.datetime.combine(datetime.date.today(), time2) + datetime.timedelta(days=1) 
        else:
            dt1 = datetime.datetime.combine(datetime.date.today(), time1)
            dt2 = datetime.datetime.combine(datetime.date.today(), time2)
            
        delta = (dt2 - dt1).total_seconds()
        hours = round(delta/3600, 2)
        
        if hours >= 7: 
            list_hour.append(hours - 1)
        else: 
            list_hour.append(hours)
    
    return list_hour

#To evenly distribute the total number of hours worked to all the studies allocated 
def crt_hours_dict(worksheet):
    study_hour_dict = {}
    crt = crt_names(worksheet)
    for key in crt:
        first_row = crt[key].row
        second_row = crt[key].offset(row=1, column=0).row
        
        sliced_worksheet = list(worksheet.iter_rows(min_row= first_row, max_row=second_row, min_col= 2, values_only= True))
        
        timing = [time_format(x) for x in sliced_worksheet[0] if x is not None and x != "-" and not isinstance(x, str)]
        timing = time_in_hours(timing)
        studies= [x for x in sliced_worksheet[1] if x is not None]
        
        for i, study in enumerate(studies):
            list_of_studies = extract_study_codes(study, worksheet)
            list_of_studies = [substring.strip() for substring in list_of_studies]
            
            for s in list_of_studies: 
                if s in study_hour_dict:
                    study_hour_dict[s] += timing[i]/(len(list_of_studies))
                else:
                    study_hour_dict[s] = timing[i]/(len(list_of_studies))
    
    return study_hour_dict

# This function is to find the row number of the keyword in the excel sheet
def row_identifier(worksheet, keyword):
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value == str(keyword):
                start_row_studies = cell.row
                
    return start_row_studies

# This function is to find the names of the studies in the excel sheet
def study_names(worksheet):
    study_names_unique = {"GS"}
    for row in worksheet.iter_rows(min_row=(row_identifier(worksheet, "Studies"))+1, max_row=row_identifier(worksheet, "AM GS")-1, min_col=0, max_col=1 ):
        for cell in row:
            if isinstance(cell.value, str):
                study_names_unique.add(cell.value.split(" ")[0])
    return study_names_unique

# This function is to extract the study codes from the study text and check if they are in the list of study names
def extract_study_codes(study_text, worksheet):
    study_keys = study_names(worksheet)
    matches = [code for code in re.findall(r'[A-Z]{2,5}', study_text.upper()) if code in study_keys]
    return matches


# Main function 
def main(files):
  excel_files = files
  total_dict = {}  
  st.write()
  if len(excel_files) > 0:
    for file_ in excel_files:
      workbook = openpyxl.load_workbook(file_)
      worksheet = workbook["Sheet1"]
      results = crt_hours_dict(worksheet)
      st.write("-"*60)
      st.write(f'''Now tabulating for this excel roster: {file_.name}''',"\n")
      for key, value in results.items():
        st.write(key, ":", value)
        total_dict[key] = total_dict.get(key, 0) + value  
      st.write()
    
    st.write()
    st.write("="*60)
    st.subheader(f'''Total all excel files''')
    for key, value in total_dict.items():
        st.write(key, ":", value)
    st.write()
    st.write(f''' We are done with a total of {len(excel_files)} excel files \U0001F601 ''')
  return 

#Initialize with Menu Bar 
selected = option_menu(
    menu_title = None, 
    options = ["Home", "Demo", "About"],
    icons = ["house-fill", "camera-video-fill", "question-circle-fill"],
    menu_icon = "cast",
    default_index = 0, 
    orientation = "horizontal",
)

# Depending on which page is chosen 
if selected == "Home":

    st.title("\U0001F4C8 CRT Hours and Study Allocated \U0001F4C9")
    st.subheader("Input Excel Files")

    st.write("\u2757 Please remember to remove date columns that are out of the month of interest first before running \u2757")

    uploaded_files = st.file_uploader("Choose the Excel Files to Upload (.xlsx)", 
                                    type ="xlsx", accept_multiple_files=True)

    if uploaded_files:
        main(uploaded_files)

elif selected == "Demo":
    st.title("Demo Video")
    video_file = open("demo.mp4", "rb")
    video_bytes = video_file.read()
    
    st.video(video_bytes)
    
elif selected == "About":
    st.title("About")
    st.write("""
            ## Note 
            1. Input are excel files in .xlsx format with dates that are not in the desired month removed first (Preprocess excel files by removing the entire column)
            2. Only CRTA are tabulated with study hours evenly distributed to studies allocated as determined by the studies at the top
            3. For shifts more than or equal to 7 hours, an hour break is mandatory and subtracted from the total hours 
            """)

    for i in range(5):
        st.write("\n")    
    
    st.write("Solely for use at Lilly Centre for Clinical Phramcology Trials @ Synapse 2026 June, Version 2.0")
    st.write("Created by Zachery Lee Wei Quan using Streamlit and hosted on Streamlit Community Cloud")
    
